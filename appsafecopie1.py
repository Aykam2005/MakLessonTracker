from flask import Flask, render_template, request, jsonify, session
import pandas as pd
import os
from openpyxl import load_workbook
import uuid
import sys

app = Flask(__name__)
app.secret_key = 'lesson_tracker_secret_key'  # Required for session management

# Configure the path to your Excel file in OneDrive
# Replace this path with your actual OneDrive path to the Excel file
# Example for Windows: "C:\\Users\\YourUsername\\OneDrive\\lessonlogtest.xlsm"
# Example for Mac: "/Users/YourUsername/OneDrive/lessonlogtest.xlsm"
EXCEL_FILE = "/Users/MakyaBertrand/OneDrive/lessonlogtest.xlsm"  # Default local path

# Check if file exists at the default path, otherwise look in OneDrive
import os
onedrive_path = os.path.expanduser("~/OneDrive/lessonlogtestcopie.xlsm")  # Common OneDrive path
if not os.path.exists(EXCEL_FILE) and os.path.exists(onedrive_path):
    EXCEL_FILE = onedrive_path
    print(f"Using Excel file from OneDrive: {EXCEL_FILE}")

# Dynamically determine the Excel file path based on OS
def get_excel_file_path():
    """Determine the correct path to the Excel file based on the operating system."""
    # Check if the Excel file is in the current directory
    current_dir_file = os.path.join(os.getcwd(), 'lessonlogtestcopie1.xlsm')
    if os.path.exists(current_dir_file):
        print(f"Found Excel file in current directory: {current_dir_file}")
        return current_dir_file
    
    # Define the fixed path for the file
    fixed_path = "/Users/makyabertrand/Library/CloudStorage/OneDrive-Personnel/LessonTrackerCopie/lessonlogtestcopie1.xlsm"
    
    # Check if the fixed path exists
    if os.path.exists(fixed_path):
        print(f"Found Excel file at fixed path: {fixed_path}")
        return fixed_path
    
    # Common OneDrive paths by operating system
    if sys.platform.startswith('win'):
        # Windows paths
        onedrive_paths = [
            os.path.join(os.path.expanduser('~'), 'OneDrive', 'lessonlogtestcopie1.xlsm'),
            os.path.join(os.path.expanduser('~'), 'OneDrive - Personnel', 'lessonlogtestcopie1.xlsm'),
            os.path.join(os.path.expanduser('~'), 'OneDrive - Personal', 'lessonlogtestcopie1.xlsm')
        ]
    else:
        # macOS/Linux paths
        onedrive_paths = [
            os.path.join(os.path.expanduser('~'), 'OneDrive', 'lessonlogtestcopie1.xlsm'),
            os.path.join(os.path.expanduser('~'), 'Library', 'CloudStorage', 'OneDrive-Personnel', 'LessonTrackerCopie', 'lessonlogtestcopie1.xlsm'),
            os.path.join(os.path.expanduser('~'), 'Library', 'CloudStorage', 'OneDrive-Personnel', 'lessonlogtestcopie1.xlsm'),
            os.path.join(os.path.expanduser('~'), 'Library', 'CloudStorage', 'OneDrive-Personal', 'lessonlogtestcopie1.xlsm')
        ]
    
    # Check all possible paths
    for path in onedrive_paths:
        if os.path.exists(path):
            print(f"Found Excel file at: {path}")
            return path
    
    # If we get here, we couldn't find the file
    print(f"WARNING: Excel file not found at any expected location.")
    print(f"Please place the 'lessonlogtestcopie1.xlsm' file in the same directory as this script.")
    
    # Return the current directory path as a fallback
    return current_dir_file

# Set the Excel file path
EXCEL_FILE = get_excel_file_path()
print(f"Using Excel file: {EXCEL_FILE}")

# Create a session-based storage for pending lessons
def get_pending_lessons():
    if 'pending_lessons' not in session:
        session['pending_lessons'] = {}
    return session['pending_lessons']

def add_pending_lesson(lesson_data):
    pending_lessons = get_pending_lessons()
    lesson_id = str(uuid.uuid4())
    lesson_data['id'] = lesson_id
    pending_lessons[lesson_id] = lesson_data
    session['pending_lessons'] = pending_lessons
    session.modified = True
    return lesson_id

def update_pending_lesson(lesson_id, lesson_data):
    pending_lessons = get_pending_lessons()
    if lesson_id in pending_lessons:
        lesson_data['id'] = lesson_id
        pending_lessons[lesson_id] = lesson_data
        session['pending_lessons'] = pending_lessons
        session.modified = True
        return True
    return False

def delete_pending_lesson(lesson_id):
    pending_lessons = get_pending_lessons()
    if lesson_id in pending_lessons:
        del pending_lessons[lesson_id]
        session['pending_lessons'] = pending_lessons
        session.modified = True
        return True
    return False

def clear_pending_lessons():
    session['pending_lessons'] = {}
    session.modified = True

def get_excel_data():
    """Read data from 'list info' and 'skater info' sheets"""
    try:
        print(f"Trying to load Excel file from: {EXCEL_FILE}")
        
        # Check if file exists
        if not os.path.exists(EXCEL_FILE):
            print(f"Error: Excel file not found at {EXCEL_FILE}")
            return {
                'athletes': [],
                'durations': [],
                'lesson_types': [],
                'focus_areas': [],
                'sheet_names': []
            }
        
        excel_file = pd.ExcelFile(EXCEL_FILE, engine='openpyxl')

        # Read from 'skater info' for athletes
        if 'skater info' in excel_file.sheet_names:
            skater_df = pd.read_excel(EXCEL_FILE, sheet_name='skater info', engine='openpyxl')
            athlete_col = next((col for col in skater_df.columns if any(name in col.lower() for name in ['name', 'athlete', 'student'])), skater_df.columns[0])
            athletes = skater_df[athlete_col].dropna().tolist()
        else:
            print("Warning: 'skater info' sheet not found")
            athletes = []

        # Read from 'list info' for durations, lesson types, focus areas
        if 'list info' in excel_file.sheet_names:
            list_df = pd.read_excel(EXCEL_FILE, sheet_name='list info', engine='openpyxl')
            durations = list_df['Durations'].dropna().tolist() if 'Durations' in list_df.columns else []
            lesson_types = list_df['Lesson Types'].dropna().tolist() if 'Lesson Types' in list_df.columns else []
            focus_areas = list_df['Focus Areas'].dropna().tolist() if 'Focus Areas' in list_df.columns else []
        else:
            print("Warning: 'list info' sheet not found")
            durations = []
            lesson_types = []
            focus_areas = []

        print(f"Successfully loaded data: {len(athletes)} athletes, {len(durations)} durations")
        return {
            'athletes': athletes,
            'durations': durations,
            'lesson_types': lesson_types,
            'focus_areas': focus_areas,
            'sheet_names': excel_file.sheet_names
        }
    except Exception as e:
        print(f"Error reading Excel file at {EXCEL_FILE}: {e}")
        return {
            'athletes': [],
            'durations': [],
            'lesson_types': [],
            'focus_areas': [],
            'sheet_names': []
        }

def add_lessons_to_excel(lessons):
    """Add multiple lesson records to the 'lesson log' sheet in the Excel file"""
    try:
        # Check if file exists
        if not os.path.exists(EXCEL_FILE):
            print(f"Error: Excel file not found at {EXCEL_FILE}")
            return False
            
        wb = load_workbook(EXCEL_FILE, keep_vba=True)
        if 'lesson log' not in wb.sheetnames:
            print("Error: 'lesson log' sheet not found.")
            return False

        sheet = wb['lesson log']

        # Get headers from the first row
        headers = [cell.value for cell in sheet[1]]
        print(f"Found headers: {headers}")

        try:
            # Create a map of expected columns
            header_map = {
                'date': headers.index('Date'),
                'athlete': headers.index("Athlete's Name"),
                'duration': headers.index('Durations'),
                'lesson_type': headers.index('Lesson Types'),
                'focus_area': headers.index('Focus Areas')
            }
        except ValueError as e:
            print(f"Error mapping headers: {e}. Headers found: {headers}")
            return False

        # Find the first empty row starting from row 2
        row = 2
        while sheet.cell(row=row, column=1).value:  # Check if Date column is not empty
            row += 1

        # Insert all lessons
        for lesson_data in lessons:
            sheet.cell(row=row, column=header_map['date'] + 1, value=lesson_data['date'])
            sheet.cell(row=row, column=header_map['athlete'] + 1, value=lesson_data['athlete'])
            sheet.cell(row=row, column=header_map['duration'] + 1, value=lesson_data['duration'])
            sheet.cell(row=row, column=header_map['lesson_type'] + 1, value=lesson_data['lesson_type'])
            sheet.cell(row=row, column=header_map['focus_area'] + 1, value=lesson_data['focus_area'])
            row += 1

        wb.save(EXCEL_FILE)
        print(f"Successfully added {len(lessons)} lessons to Excel file")
        return True
    except Exception as e:
        print(f"Error adding lessons to Excel: {e}")
        return False
       
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/form-data')
def form_data():
    data = get_excel_data()
    # Include pending lessons in the response
    data['pending_lessons'] = list(get_pending_lessons().values())
    return jsonify(data)

@app.route('/submit-lesson', methods=['POST'])
def submit_lesson():
    if request.method == 'POST':
        lesson_data = {
            'date': request.form['date'],
            'athlete': request.form['athlete'],
            'duration': request.form['duration'],
            'lesson_type': request.form['lesson_type'],
            'focus_area': request.form['focus_area']
        }
        
        # Add to pending lessons
        lesson_id = add_pending_lesson(lesson_data)
        
        return jsonify({
            'status': 'success', 
            'message': 'Lesson added to pending list!',
            'lesson_id': lesson_id,
            'pending_lessons': list(get_pending_lessons().values())
        })

@app.route('/update-lesson', methods=['POST'])
def update_lesson():
    if request.method == 'POST':
        lesson_id = request.form['id']
        lesson_data = {
            'date': request.form['date'],
            'athlete': request.form['athlete'],
            'duration': request.form['duration'],
            'lesson_type': request.form['lesson_type'],
            'focus_area': request.form['focus_area']
        }
        
        success = update_pending_lesson(lesson_id, lesson_data)
        
        if success:
            return jsonify({
                'status': 'success', 
                'message': 'Lesson updated successfully!',
                'pending_lessons': list(get_pending_lessons().values())
            })
        else:
            return jsonify({
                'status': 'error', 
                'message': 'Failed to update lesson.',
                'pending_lessons': list(get_pending_lessons().values())
            })

@app.route('/delete-lesson', methods=['POST'])
def delete_lesson():
    if request.method == 'POST':
        lesson_id = request.form['id']
        
        success = delete_pending_lesson(lesson_id)
        
        if success:
            return jsonify({
                'status': 'success', 
                'message': 'Lesson deleted successfully!',
                'pending_lessons': list(get_pending_lessons().values())
            })
        else:
            return jsonify({
                'status': 'error', 
                'message': 'Failed to delete lesson.',
                'pending_lessons': list(get_pending_lessons().values())
            })

@app.route('/commit-lessons', methods=['POST'])
def commit_lessons():
    if request.method == 'POST':
        pending_lessons = list(get_pending_lessons().values())
        
        if not pending_lessons:
            return jsonify({
                'status': 'error', 
                'message': 'No pending lessons to commit.'
            })
        
        success = add_lessons_to_excel(pending_lessons)
        
        if success:
            # Clear pending lessons after successful commit
            clear_pending_lessons()
            return jsonify({
                'status': 'success', 
                'message': f'{len(pending_lessons)} lessons added to Excel successfully!'
            })
        else:
            return jsonify({
                'status': 'error', 
                'message': 'Failed to add lessons to Excel file.'
            })

if __name__ == '__main__':
    # Ensure the templates directory exists
    if not os.path.exists('templates'):
        os.makedirs('templates')
        print("Created templates directory")
    
    # Create a basic HTML template if it doesn't exist
    template_path = os.path.join('templates', 'index.html')
    if not os.path.exists(template_path):
        with open(template_path, 'w') as f:
            f.write('''<!DOCTYPE html>
<html>
<head>
    <title>Lesson Tracker</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        body { font-family: Arial, sans-serif; margin: 0; padding: 20px; }
        .container { max-width: 800px; margin: 0 auto; }
        .form-group { margin-bottom: 15px; }
        label { display: block; margin-bottom: 5px; }
        input, select { width: 100%; padding: 8px; box-sizing: border-box; }
        button { background: #4CAF50; color: white; padding: 10px 15px; border: none; cursor: pointer; }
        button:hover { background: #45a049; }
        .pending-lessons { margin-top: 30px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { padding: 8px; text-align: left; border-bottom: 1px solid #ddd; }
        .actions { display: flex; gap: 5px; }
        .error { color: red; }
        .success { color: green; }
        .file-status { margin: 20px 0; padding: 10px; border-radius: 5px; }
        .file-status.success { background-color: #d4edda; border: 1px solid #c3e6cb; }
        .file-status.error { background-color: #f8d7da; border: 1px solid #f5c6cb; }
        .file-input { margin: 20px 0; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Lesson Tracker</h1>
        <div id="message"></div>
        
        <div id="fileStatus" class="file-status">
            Checking Excel file status...
        </div>
        
        <div class="file-input">
            <h3>Excel File Location</h3>
            <p>If your Excel file is not being found automatically, enter the full path to your lessonlogtestcopie1.xlsm file:</p>
            <input type="text" id="excelFilePath" placeholder="/full/path/to/lessonlogtestcopie1.xlsm">
            <button id="updatePathButton">Update Path</button>
        </div>
        
        <form id="lessonForm">
            <div class="form-group">
                <label for="date">Date:</label>
                <input type="date" id="date" name="date" required>
            </div>
            
            <div class="form-group">
                <label for="athlete">Athlete:</label>
                <select id="athlete" name="athlete" required>
                    <option value="">Select an athlete</option>
                </select>
            </div>
            
            <div class="form-group">
                <label for="duration">Duration:</label>
                <select id="duration" name="duration" required>
                    <option value="">Select duration</option>
                </select>
            </div>
            
            <div class="form-group">
                <label for="lesson_type">Lesson Type:</label>
                <select id="lesson_type" name="lesson_type" required>
                    <option value="">Select lesson type</option>
                </select>
            </div>
            
            <div class="form-group">
                <label for="focus_area">Focus Area:</label>
                <select id="focus_area" name="focus_area" required>
                    <option value="">Select focus area</option>
                </select>
            </div>
            
            <button type="submit">Add Lesson</button>
        </form>
        
        <div class="pending-lessons">
            <h2>Pending Lessons</h2>
            <table id="pendingLessonsTable">
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Athlete</th>
                        <th>Duration</th>
                        <th>Lesson Type</th>
                        <th>Focus Area</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    <!-- Pending lessons will be displayed here -->
                </tbody>
            </table>
            
            <div style="margin-top: 20px;">
                <button id="commitButton">Save All Lessons</button>
            </div>
        </div>
    </div>

    <script>
        // Fetch form data from the server
        async function fetchFormData() {
            try {
                const response = await fetch('/form-data');
                const data = await response.json();
                
                // Update Excel file status
                updateFileStatus(data);
                
                // Populate dropdowns
                populateDropdown('athlete', data.athletes);
                populateDropdown('duration', data.durations);
                populateDropdown('lesson_type', data.lesson_types);
                populateDropdown('focus_area', data.focus_areas);
                
                // Display pending lessons
                displayPendingLessons(data.pending_lessons || []);
            } catch (error) {
                showMessage('Error loading form data: ' + error.message, 'error');
                
                // Update file status to error
                const fileStatus = document.getElementById('fileStatus');
                fileStatus.textContent = 'Error connecting to Excel file. Please check the file path.';
                fileStatus.className = 'file-status error';
            }
        }
        
        // Update the Excel file status display
        function updateFileStatus(data) {
            const fileStatus = document.getElementById('fileStatus');
            
            if (data.athletes.length > 0 || data.durations.length > 0) {
                fileStatus.textContent = 'Successfully connected to Excel file!';
                fileStatus.className = 'file-status success';
            } else {
                fileStatus.textContent = 'Could not find or read data from Excel file. Please check the file path.';
                fileStatus.className = 'file-status error';
            }
        }
        
        // Populate a dropdown with options
        function populateDropdown(id, options) {
            const dropdown = document.getElementById(id);
            
            // Keep the first option (placeholder)
            const placeholder = dropdown.options[0];
            dropdown.innerHTML = '';
            dropdown.appendChild(placeholder);
            
            // Add options from the data
            options.forEach(option => {
                const optionElement = document.createElement('option');
                optionElement.value = option;
                optionElement.textContent = option;
                dropdown.appendChild(optionElement);
            });
        }
        
        // Display pending lessons in the table
        function displayPendingLessons(lessons) {
            const tbody = document.querySelector('#pendingLessonsTable tbody');
            tbody.innerHTML = '';
            
            if (lessons.length === 0) {
                const row = document.createElement('tr');
                row.innerHTML = '<td colspan="6">No pending lessons</td>';
                tbody.appendChild(row);
                return;
            }
            
            lessons.forEach(lesson => {
                const row = document.createElement('tr');
                row.innerHTML = `
                    <td>${lesson.date}</td>
                    <td>${lesson.athlete}</td>
                    <td>${lesson.duration}</td>
                    <td>${lesson.lesson_type}</td>
                    <td>${lesson.focus_area}</td>
                    <td class="actions">
                        <button class="edit-button" data-id="${lesson.id}">Edit</button>
                        <button class="delete-button" data-id="${lesson.id}">Delete</button>
                    </td>
                `;
                tbody.appendChild(row);
            });
            
            // Add event listeners to edit and delete buttons
            document.querySelectorAll('.edit-button').forEach(button => {
                button.addEventListener('click', () => editLesson(button.getAttribute('data-id')));
            });
            
            document.querySelectorAll('.delete-button').forEach(button => {
                button.addEventListener('click', () => deleteLesson(button.getAttribute('data-id')));
            });
        }
        
        // Show a message to the user
        function showMessage(message, type = 'success') {
            const messageElement = document.getElementById('message');
            messageElement.textContent = message;
            messageElement.className = type;
            
            // Clear message after 5 seconds
            setTimeout(() => {
                messageElement.textContent = '';
                messageElement.className = '';
            }, 5001);
        }
        
        // Submit the lesson form
        document.getElementById('lessonForm').addEventListener('submit', async (event) => {
            event.preventDefault();
            
            const formData = new FormData(event.target);
            let editId = document.getElementById('lessonForm').getAttribute('data-edit-id');
            
            try {
                let url = '/submit-lesson';
                
                // If we're editing an existing lesson, use the update endpoint
                if (editId) {
                    url = '/update-lesson';
                    formData.append('id', editId);
                }
                
                const response = await fetch(url, {
                    method: 'POST',
                    body: formData
                });
                
                const result = await response.json();
                
                if (result.status === 'success') {
                    showMessage(result.message);
                    // Reset form
                    document.getElementById('lessonForm').reset();
                    document.getElementById('lessonForm').removeAttribute('data-edit-id');
                    document.querySelector('button[type="submit"]').textContent = 'Add Lesson';
                    
                    // Update pending lessons display
                    displayPendingLessons(result.pending_lessons);
                } else {
                    showMessage(result.message, 'error');
                }
            } catch (error) {
                showMessage('Error submitting form: ' + error.message, 'error');
            }
        });
        
        // Edit a lesson
        async function editLesson(id) {
            try {
                const response = await fetch('/form-data');
                const data = await response.json();
                
                // Find the lesson to edit
                const lesson = data.pending_lessons.find(l => l.id === id);
                
                if (lesson) {
                    // Fill the form with lesson data
                    document.getElementById('date').value = lesson.date;
                    document.getElementById('athlete').value = lesson.athlete;
                    document.getElementById('duration').value = lesson.duration;
                    document.getElementById('lesson_type').value = lesson.lesson_type;
                    document.getElementById('focus_area').value = lesson.focus_area;
                    
                    // Set form to edit mode
                    document.getElementById('lessonForm').setAttribute('data-edit-id', id);
                    document.querySelector('button[type="submit"]').textContent = 'Update Lesson';
                    
                    // Scroll to the form
                    document.getElementById('lessonForm').scrollIntoView({ behavior: 'smooth' });
                }
            } catch (error) {
                showMessage('Error loading lesson data: ' + error.message, 'error');
            }
        }
        
        // Delete a lesson
        async function deleteLesson(id) {
            if (confirm('Are you sure you want to delete this lesson?')) {
                try {
                    const formData = new FormData();
                    formData.append('id', id);
                    
                    const response = await fetch('/delete-lesson', {
                        method: 'POST',
                        body: formData
                    });
                    
                    const result = await response.json();
                    
                    if (result.status === 'success') {
                        showMessage(result.message);
                        // Update pending lessons display
                        displayPendingLessons(result.pending_lessons);
                    } else {
                        showMessage(result.message, 'error');
                    }
                } catch (error) {
                    showMessage('Error deleting lesson: ' + error.message, 'error');
                }
            }
        }
        
        // Commit all pending lessons
        document.getElementById('commitButton').addEventListener('click', async () => {
            try {
                const response = await fetch('/commit-lessons', {
                    method: 'POST'
                });
                
                const result = await response.json();
                
                if (result.status === 'success') {
                    showMessage(result.message);
                    // Refresh pending lessons (should be empty now)
                    fetchFormData();
                } else {
                    showMessage(result.message, 'error');
                }
            } catch (error) {
                showMessage('Error committing lessons: ' + error.message, 'error');
            }
        });
        
        // Initial load
        document.addEventListener('DOMContentLoaded', fetchFormData);
    </script>
</body>
</html>''')
        print(f"Created basic HTML template at {template_path}")
    
    # Start the Flask app
    app.run(host='0.0.0.0', port=5001, debug=True)
